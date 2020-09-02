import openpyxl

# Import sheet and setup so that I can work with it
importSheet = "C:\Users\esswe\Documents\Coding\MyProjects\BwlMcSheetScraper\import.py"
wbImport = openpyxl.load_workbook(importSheet)
sheetImport = wbImport.active

# Import sheet and setup so that I can work with it
exportSheet = "C:\Users\esswe\Documents\Coding\MyProjects\BwlMcSheetScraper\cnwe.py"
wbExport = openpyxl.load_workbook(exportSheet)
sheetExport = wbExport.active

rowCurrent = 1
for i in range(1,151):
    colCurrent = 2
    while(1):
        if(sheetImport.cell(row=i,column=colCurrent) is not None and type(sheetImport.cell(row=i, column=colCurrent).value == str)):
            
        else:
            break